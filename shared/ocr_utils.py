"""
RAG Medan v3 - Shared OCR Utils
Utilities untuk OCR dan text extraction dari berbagai format file.
Mendukung dua mode:
  - local: PaddleOCR (default)
  - api  : LLM via Router API (9router)
"""
import os
import re
import time
import tempfile
import logging
import hashlib
from copy import deepcopy
from typing import Any, Callable, Dict, List, Optional

from config import config
from shared.pdf_layout_extractor import PdfLayoutResult, extract_pdf_layout

logger = logging.getLogger("ocr_utils")

# Lazy-loaded PaddleOCR engine (singleton)
_ocr_engine: Optional[object] = None
_pdf_layout_cache: Dict[tuple, PdfLayoutResult] = {}


def _notify_progress(
    progress_callback: Optional[Callable[..., None]] = None,
    **payload: Any,
) -> None:
    """Invoke progress callback without breaking extraction flow."""
    if not progress_callback:
        return
    try:
        progress_callback(**payload)
    except TypeError:
        try:
            progress_callback(payload)
        except Exception:
            logger.debug("[OCR] progress callback rejected payload", exc_info=True)
    except Exception:
        logger.debug("[OCR] progress callback failed", exc_info=True)


def get_ocr_engine():
    """Lazy load PaddleOCR engine (singleton)."""
    global _ocr_engine
    if _ocr_engine is None:
        logger.info("[OCR] Initializing PaddleOCR engine...")
        from paddleocr import PaddleOCR
        _ocr_engine = PaddleOCR(lang="id", use_angle_cls=True)
        logger.info("[OCR] PaddleOCR engine ready")
    return _ocr_engine


def _ocr_image_bytes(img_bytes: bytes) -> str:
    """OCR dari bytes gambar.

    Routing otomatis berdasarkan config.OCR_MODE:
    - 'api'  : kirim ke LLM Router API (menghasilkan Markdown)
    - 'local': gunakan PaddleOCR (menghasilkan plain text)
    """
    if config.OCR_MODE == "api":
        from shared.llm_ocr import call_llm_ocr
        logger.info("[OCR] Mode API: mengirim gambar ke LLM Router.")
        result = call_llm_ocr(img_bytes)
        if not result:
            logger.warning("[OCR] Mode API: hasil kosong dari LLM Router (gambar mungkin tidak terbaca).")
        return result

    # --- Mode local: PaddleOCR ---
    temp_file_path = None
    try:
        with tempfile.NamedTemporaryFile(suffix=".png", delete=False) as temp_file:
            temp_file.write(img_bytes)
            temp_file.flush()
            temp_file_path = temp_file.name

        ocr_engine = get_ocr_engine()
        ocr_result = ocr_engine.ocr(temp_file_path)
        if ocr_result and len(ocr_result) > 0 and ocr_result[0]:
            return "\n".join([line[1][0] for line in ocr_result[0] if line and len(line) > 1])
        return ""
    except Exception as e:
        logger.warning(f"[OCR] Gagal OCR image: {e}")
        return ""
    finally:
        if temp_file_path and os.path.exists(temp_file_path):
            try:
                os.remove(temp_file_path)
            except Exception:
                pass


def _clean_page_text(page_text: str) -> str:
    """Bersihkan header/footer, spasi dobel, nomor halaman."""
    if not page_text:
        return ""
    cleaned_text = re.sub(r"^\s*\d{1,4}\s*$", "", page_text, flags=re.MULTILINE)
    cleaned_text = re.sub(r"[ \t]+", " ", cleaned_text)
    cleaned_text = re.sub(r"\n{3,}", "\n\n", cleaned_text)
    return cleaned_text.strip()


def clean_ocr_text(text: str) -> str:
    """Membersihkan hasil OCR dari formatting berlebihan."""
    if not text:
        return ""

    text = re.sub(r'\n{3,}', '\n\n', text)
    lines = [line.strip() for line in text.split('\n')]

    cleaned_lines = []
    prev_line = None
    for line in lines:
        if line and line != prev_line:
            cleaned_lines.append(line)
            prev_line = line
        elif not line and prev_line:
            cleaned_lines.append('')
            prev_line = None

    text = '\n'.join(cleaned_lines)
    text = re.sub(r' {2,}', ' ', text)
    text = re.sub(r'([a-z,])\n([a-z])', r'\1 \2', text)

    return text.strip()


def format_for_display(text: str) -> str:
    """Format text untuk display dengan paragraph breaks yang jelas."""
    if not text:
        return ""
    
    paragraphs = re.split(r'\n\n+', text)
    cleaned_paragraphs = []
    
    for para in paragraphs:
        para = para.strip()
        if para:
            para = para.replace('\n', ' ')
            para = re.sub(r' {2,}', ' ', para)
            cleaned_paragraphs.append(para)
    
    return '\n\n'.join(cleaned_paragraphs)


def calculate_file_hash(file_path: str) -> str:
    """Hitung SHA256 hash dari file."""
    sha256_hash = hashlib.sha256()
    with open(file_path, "rb") as f:
        for chunk in iter(lambda: f.read(8192), b""):
            sha256_hash.update(chunk)
    return sha256_hash.hexdigest()


def calculate_content_hash(text: str) -> str:
    """Hitung SHA256 hash dari content text."""
    normalized = re.sub(r'\s+', ' ', text.strip().lower())
    return hashlib.sha256(normalized.encode('utf-8')).hexdigest()


def _ocr_pdf_page(
    page_number: int,
    page,
    *,
    total_page_count: int,
    dpi: int,
    retry_dpi: int,
    progress_callback: Optional[Callable[..., None]] = None,
) -> str:
    """Render satu halaman PDF menjadi gambar lalu kirim ke OCR engine.

    Mode 'api': kirim PNG bytes ke LLM Router API. Tidak ada retry-DPI
    karena LLM dapat memproses gambar dengan resolusi apapun; kualitas
    output tergantung pada akurasi model.
    Mode 'local': gunakan PaddleOCR dengan retry-DPI jika teks < 20 char.
    """
    _notify_progress(
        progress_callback,
        stage="ocr",
        message=f"OCR halaman {page_number}/{total_page_count}...",
        total_pages=total_page_count,
        current_page=page_number,
        used_ocr=True,
    )
    page_pixmap = page.get_pixmap(dpi=dpi)
    image_bytes = page_pixmap.tobytes("png")

    # Mode API: tidak ada retry-DPI; jeda ditangani oleh call_llm_ocr.
    # Cukup delegasikan ke _ocr_image_bytes yang sudah menghandle routing.
    if config.OCR_MODE == "api":
        result_text = _ocr_image_bytes(image_bytes)
        char_count = len(result_text)
        if char_count < 20:
            logger.warning(
                f"[PDF][LLM-OCR] Halaman {page_number}/{total_page_count}: "
                f"hasil sangat pendek ({char_count} karakter). "
                "Halaman mungkin kosong, gambar buram, atau LLM gagal mengenali teks."
            )
        else:
            logger.info(
                f"[PDF][LLM-OCR] Halaman {page_number}/{total_page_count} selesai: "
                f"{char_count} karakter diekstrak."
            )
        _notify_progress(
            progress_callback,
            stage="ocr",
            message=f"OCR halaman {page_number}/{total_page_count} selesai (API, {char_count} chars).",
            total_pages=total_page_count,
            current_page=page_number,
            used_ocr=True,
        )
        return result_text

    # Mode local: PaddleOCR dengan retry-DPI
    cleaned_text = _clean_page_text(_ocr_image_bytes(image_bytes))

    if len(cleaned_text) < 20 and retry_dpi > dpi:
        logger.info(f"[PDF] Retry OCR halaman {page_number} dengan dpi={retry_dpi}")
        _notify_progress(
            progress_callback,
            stage="ocr",
            message=f"Retry OCR halaman {page_number}/{total_page_count} dengan kualitas lebih tinggi...",
            total_pages=total_page_count,
            current_page=page_number,
            used_ocr=True,
        )
        retry_pixmap = page.get_pixmap(dpi=retry_dpi)
        retry_bytes = retry_pixmap.tobytes("png")
        retry_text = _clean_page_text(_ocr_image_bytes(retry_bytes))
        if len(retry_text) > len(cleaned_text):
            cleaned_text = retry_text

    _notify_progress(
        progress_callback,
        stage="ocr",
        message=f"OCR halaman {page_number}/{total_page_count} selesai.",
        total_pages=total_page_count,
        current_page=page_number,
        used_ocr=True,
    )
    return cleaned_text


def _extract_pdf_layout_result(
    pdf_path: str,
    dpi: int = 150,
    retry_dpi: int = 200,
    progress_callback: Optional[Callable[..., None]] = None,
) -> PdfLayoutResult:
    """Ekstrak teks dan block PDF dengan layout-aware text layer + OCR fallback."""
    import fitz

    try:
        stat = os.stat(pdf_path)
        cache_key = (os.path.abspath(pdf_path), stat.st_mtime_ns, stat.st_size, dpi, retry_dpi)
        cached = _pdf_layout_cache.get(cache_key)
        if cached is not None:
            return deepcopy(cached)
    except OSError:
        cache_key = None

    logger.info(f"[PDF] Opening PDF: {pdf_path}")
    try:
        pdf_document = fitz.open(pdf_path)
    except Exception as e:
        logger.error(f"[PDF] Failed to open PDF: {e}")
        raise

    total_page_count = len(pdf_document)
    pdf_document.close()
    logger.info(f"[PDF] Total pages: {total_page_count}")
    _notify_progress(
        progress_callback,
        stage="extracting",
        message=f"PDF terdeteksi {total_page_count} halaman. Memulai ekstraksi layout-aware...",
        total_pages=total_page_count,
        current_page=0,
    )

    ocr_used_pages: set[int] = set()

    def ocr_callback(page_number: int, page) -> str:
        try:
            ocr_used_pages.add(page_number)
            return _ocr_pdf_page(
                page_number,
                page,
                total_page_count=total_page_count,
                dpi=dpi,
                retry_dpi=retry_dpi,
                progress_callback=progress_callback,
            )
        except Exception as e:
            logger.warning(f"[PDF] Gagal render/OCR halaman {page_number}: {e}")
            _notify_progress(
                progress_callback,
                stage="ocr",
                message=f"OCR halaman {page_number}/{total_page_count} gagal, melanjutkan ke halaman berikutnya.",
                total_pages=total_page_count,
                current_page=page_number,
                used_ocr=True,
            )
            return ""

    result = extract_pdf_layout(
        pdf_path,
        source_kind="ocr",
        ocr_page_callback=ocr_callback,
    )
    for page_number in range(1, total_page_count + 1):
        _notify_progress(
            progress_callback,
            stage="extracting",
            message=f"Ekstraksi layout halaman {page_number}/{total_page_count} selesai.",
            total_pages=total_page_count,
            current_page=page_number,
            used_ocr=page_number in ocr_used_pages,
        )
    if cache_key is not None:
        _pdf_layout_cache.clear()
        _pdf_layout_cache[cache_key] = deepcopy(result)
    return result


def _extract_pdf_pages(
    pdf_path: str,
    dpi: int = 150,
    retry_dpi: int = 200,
    progress_callback: Optional[Callable[..., None]] = None,
) -> Dict[int, str]:
    """Ekstrak teks PDF per halaman dengan layout-aware hybrid extraction."""
    result = _extract_pdf_layout_result(
        pdf_path,
        dpi=dpi,
        retry_dpi=retry_dpi,
        progress_callback=progress_callback,
    )
    return dict(sorted(result.pages.items(), key=lambda x: x[0]))


def extract_text_from_file(
    file_path: str,
    lang: str = "id",
    return_pages: bool = False,
    progress_callback: Optional[Callable[..., None]] = None,
):
    """
    Ekstraksi teks dari berbagai format file.
    Mendukung: PDF, DOCX, XLSX, TXT, dan gambar (JPG, PNG).
    
    Args:
        file_path: Path ke file
        lang: Bahasa untuk OCR (default: "id")
        return_pages: Jika True return dict, False return string gabungan
    """
    file_extension = os.path.splitext(file_path)[1].lower()
    extracted_pages = {}

    if file_extension == ".pdf":
        extracted_pages = _extract_pdf_pages(
            file_path,
            dpi=config.OCR_PDF_DPI,
            retry_dpi=config.OCR_PDF_DPI_RETRY,
            progress_callback=progress_callback,
        )

    elif file_extension in [".jpg", ".jpeg", ".png"]:
        _notify_progress(progress_callback, stage="ocr", message="Memulai OCR gambar...", current_page=1, total_pages=1)
        try:
            with open(file_path, "rb") as img_f:
                img_bytes = img_f.read()
            extracted_text = _ocr_image_bytes(img_bytes)
        except Exception as e:
            logger.warning(f"[OCR] Gagal OCR image file {file_path}: {e}")
            extracted_text = ""
        # Mode local: teks plain, perlu dibersihkan
        # Mode api: output Markdown, _clean_page_text aman untuk Markdown
        extracted_pages[1] = _clean_page_text(extracted_text) if config.OCR_MODE == "local" else extracted_text
        _notify_progress(progress_callback, stage="ocr", message="OCR gambar selesai.", current_page=1, total_pages=1)

    elif file_extension == ".docx":
        try:
            _notify_progress(progress_callback, stage="extracting", message="Membaca dokumen DOCX...", current_page=1, total_pages=1)
            # Gunakan _extract_docx_blocks (yang sudah mendukung tabel) untuk
            # memastikan content_hash mencerminkan seluruh isi dokumen.
            docx_blocks = _extract_docx_blocks(file_path)
        except Exception as e:
            logger.warning(f"[DOCX] Gagal membuka file DOCX {file_path}: {e}")
            extracted_pages[1] = ""
            if return_pages:
                return extracted_pages
            return ""

        text_parts = []
        for block in docx_blocks:
            block_text = block.get("text", "").strip()
            if not block_text:
                continue
            # Tambah separator antar heading agar batas chunk terlihat jelas
            if block.get("block_type") == "heading":
                text_parts.append("\n\n" + block_text)
            else:
                text_parts.append(block_text)
        extracted_text = "\n".join(text_parts)
        extracted_pages[1] = _clean_page_text(extracted_text)
        _notify_progress(progress_callback, stage="extracting", message="Ekstraksi DOCX selesai.", current_page=1, total_pages=1)

    elif file_extension == ".xlsx":
        import openpyxl

        try:
            _notify_progress(progress_callback, stage="extracting", message="Membaca workbook XLSX...", current_page=1, total_pages=1)
            # read_only=True: streaming mode — RAM konstan berapapun ukuran file
            workbook = openpyxl.load_workbook(file_path, data_only=True, read_only=True)
            all_sheets_text = []
            for sheet in workbook.worksheets:
                sheet_text = f"=== Sheet: {sheet.title} ===\n"
                for row in sheet.iter_rows(values_only=True):
                    row_text = " | ".join([str(cell) if cell is not None else "" for cell in row])
                    if row_text.strip():
                        sheet_text += row_text + "\n"
                all_sheets_text.append(sheet_text)
            extracted_text = "\n\n".join(all_sheets_text)
            extracted_pages[1] = _clean_page_text(extracted_text)
            _notify_progress(progress_callback, stage="extracting", message="Ekstraksi XLSX selesai.", current_page=1, total_pages=1)
        except Exception as e:
            logger.warning(f"[XLSX] Gagal ekstraksi Excel file {file_path}: {e}")
            extracted_pages[1] = ""

    elif file_extension == ".txt":
        try:
            _notify_progress(progress_callback, stage="extracting", message="Membaca file TXT...", current_page=1, total_pages=1)
            with open(file_path, "r", encoding="utf-8") as txt_file:
                extracted_text = txt_file.read()
            extracted_pages[1] = _clean_page_text(extracted_text)
            _notify_progress(progress_callback, stage="extracting", message="Ekstraksi TXT selesai.", current_page=1, total_pages=1)
        except UnicodeDecodeError:
            try:
                with open(file_path, "r", encoding="latin-1") as txt_file:
                    extracted_text = txt_file.read()
                extracted_pages[1] = _clean_page_text(extracted_text)
            except Exception as e:
                logger.warning(f"[TXT] Gagal ekstraksi TXT file {file_path}: {e}")
                extracted_pages[1] = ""

    else:
        raise ValueError(f"Format file {file_extension} belum didukung untuk OCR.")

    if return_pages:
        return dict(sorted(extracted_pages.items(), key=lambda x: x[0]))

    return "\n\n".join(extracted_pages.values()).strip()


_HEADING_HINT_RE = re.compile(
    r"^(bab|bagian|section|pasal|lampiran|judul|ketentuan|persyaratan|prosedur|alur|tahapan)\b",
    re.IGNORECASE,
)
_LIST_HINT_RE = re.compile(r"^(\d+[\.\)]|[-*]|[a-zA-Z][\.\)])\s+")


def _is_heading_candidate(text: str) -> bool:
    cleaned = (text or "").strip()
    if not cleaned:
        return False
    if len(cleaned) > 120:
        return False
    if _HEADING_HINT_RE.match(cleaned):
        return True
    if cleaned.endswith((".", "?", "!", ";", ":")):
        return False
    if cleaned.isupper() and len(cleaned.split()) <= 10:
        return True
    words = cleaned.split()
    titled_words = sum(1 for word in words if word[:1].isupper())
    return len(words) <= 8 and titled_words == len(words)


def _heading_level_from_text(text: str) -> int:
    lowered = (text or "").strip().lower()
    if lowered.startswith("bab"):
        return 1
    if lowered.startswith("bagian"):
        return 2
    if lowered.startswith("pasal"):
        return 3
    return 2


def _build_plaintext_blocks(
    text: str,
    *,
    page_number: int,
    source_kind: str,
    heading_path: Optional[list[str]] = None,
    block_order_start: int = 0,
) -> list[dict]:
    blocks: list[dict] = []
    active_heading_path = list(heading_path or [])
    paragraphs = [part.strip() for part in re.split(r"\n\s*\n+", text or "") if part.strip()]
    block_order = block_order_start

    for paragraph in paragraphs:
        lines = [line.strip() for line in paragraph.splitlines() if line.strip()]
        if not lines:
            continue

        first_line = lines[0]
        block_type = "paragraph"
        heading_level = None
        heading_text = active_heading_path[-1] if active_heading_path else ""

        if _is_heading_candidate(first_line):
            block_type = "heading"
            heading_level = _heading_level_from_text(first_line)
            active_heading_path = active_heading_path[:heading_level - 1] + [first_line]
            heading_text = first_line
        elif _LIST_HINT_RE.match(first_line):
            block_type = "list_item"

        normalized = _clean_page_text(paragraph)
        if not normalized:
            continue

        blocks.append({
            "page_number": page_number,
            "text": normalized,
            "block_type": block_type,
            "heading_level": heading_level,
            "heading_text": heading_text,
            "heading_path": list(active_heading_path),
            "source_kind": source_kind,
            "block_order": block_order,
            "metadata": {},
        })
        block_order += 1

    return blocks


def _blocks_from_markdown(
    text: str,
    *,
    page_number: int,
    source_kind: str = "ocr",
    block_order_start: int = 0,
) -> List[dict]:
    """Parse teks Markdown hasil LLM OCR menjadi dictionary block standar.

    Menghasilkan struktur yang IDENTIK dengan output ekstraksi lokal sehingga
    pipeline chunking & embedding tidak perlu diubah.

    Elemen yang dikenali:
    - Heading  : baris dimulai dengan # s/d ######
    - Tabel    : sekelompok baris dengan format | ... | ... |
    - Gambar   : ![...](...)  → paragraph pendek
    - Lainnya  : paragraph / list_item
    """
    blocks: List[dict] = []
    heading_path: List[str] = []
    block_order = block_order_start

    lines = (text or "").splitlines()
    i = 0
    n = len(lines)

    # --- regex helpers ---
    _heading_re = re.compile(r"^(#{1,6})\s+(.+)$")
    _table_border_re = re.compile(r"^\|?\s*:?-{2,}:?\s*(\|\s*:?-{2,}:?\s*)+\|?\s*$")
    _table_row_re = re.compile(r"^\|.+\|\s*$")
    _list_re = re.compile(r"^(\d+[\.\)]|[-*]|[a-zA-Z][\.\)])\s+")

    def _make_block(
        btype: str,
        btext: str,
        hlevel: Optional[int],
        htext: str,
        hpath: List[str],
        skind: str,
    ) -> dict:
        return {
            "page_number": page_number,
            "text": btext.strip(),
            "block_type": btype,
            "heading_level": hlevel,
            "heading_text": htext,
            "heading_path": list(hpath),
            "source_kind": skind,
            "block_order": 0,  # akan di-assign ulang di bawah
            "metadata": {},
        }

    while i < n:
        raw_line = lines[i]
        stripped = raw_line.strip()

        # --- Lewati baris kosong ---
        if not stripped:
            i += 1
            continue

        # --- Heading ---
        m_heading = _heading_re.match(stripped)
        if m_heading:
            level = len(m_heading.group(1))
            heading_text = m_heading.group(2).strip()
            heading_path = heading_path[: level - 1] + [heading_text]
            blk = _make_block(
                "heading", heading_text, level, heading_text, heading_path, source_kind
            )
            blocks.append(blk)
            i += 1
            continue

        # --- Tabel Markdown ---
        if _table_row_re.match(stripped):
            table_lines: List[str] = []
            while i < n and (_table_row_re.match(lines[i].strip()) or _table_border_re.match(lines[i].strip())):
                table_lines.append(lines[i].strip())
                i += 1

            # Pisahkan header, border, dan baris data
            header_row: List[str] = []
            data_rows: List[List[str]] = []
            for tl in table_lines:
                if _table_border_re.match(tl):
                    continue  # Abaikan baris pemisah
                cells = [c.strip() for c in tl.strip("|").split("|")]
                if not header_row:
                    header_row = cells
                else:
                    data_rows.append(cells)

            if not header_row:
                continue

            header_text = " | ".join(header_row)
            current_heading_text = heading_path[-1] if heading_path else ""

            if not data_rows:
                # Tabel hanya header
                blk = _make_block("table_row", header_text, None, current_heading_text, heading_path, "table")
                blk["metadata"] = {"is_header": True}
                blocks.append(blk)
            else:
                for data_cells in data_rows:
                    if not any(c for c in data_cells):
                        continue
                    row_text = " | ".join(data_cells)
                    combined = f"{header_text}\n{row_text}"
                    blk = _make_block("table_row", combined, None, current_heading_text, heading_path, "table")
                    blk["metadata"] = {"header_row": header_row}
                    blocks.append(blk)
            continue

        # --- Paragraph / List Item ---
        paragraph_lines = []
        while i < n and lines[i].strip() and not _heading_re.match(lines[i].strip()) and not _table_row_re.match(lines[i].strip()):
            paragraph_lines.append(lines[i].strip())
            i += 1
        paragraph_text = " ".join(paragraph_lines).strip()
        if not paragraph_text:
            continue

        btype = "list_item" if _list_re.match(paragraph_text) else "paragraph"
        current_heading_text = heading_path[-1] if heading_path else ""
        blk = _make_block(btype, paragraph_text, None, current_heading_text, heading_path, source_kind)
        blocks.append(blk)

    # Assign block_order secara berurutan
    for idx, blk in enumerate(blocks):
        blk["block_order"] = block_order_start + idx

    return blocks


def build_blocks_from_extracted_pages(
    extracted_pages: Dict[int, str],
    *,
    source_kind: str = "ocr",
) -> list[dict]:
    """Build structured blocks from already extracted page text.

    Secara otomatis memilih parser berdasarkan config.OCR_MODE:
    - 'api'  : _blocks_from_markdown (output LLM)
    - 'local': _build_plaintext_blocks (PaddleOCR plain text)
    """
    blocks: list[dict] = []
    block_order = 0
    for page_number, page_text in sorted(extracted_pages.items(), key=lambda item: item[0]):
        if config.OCR_MODE == "api":
            page_blocks = _blocks_from_markdown(
                page_text,
                page_number=page_number,
                source_kind=source_kind,
                block_order_start=block_order,
            )
        else:
            page_blocks = _build_plaintext_blocks(
                page_text,
                page_number=page_number,
                source_kind=source_kind,
                block_order_start=block_order,
            )
        blocks.extend(page_blocks)
        block_order += len(page_blocks)
    return blocks


def _extract_docx_blocks(file_path: str) -> list[dict]:
    from docx import Document
    from docx.text.paragraph import Paragraph as DocxParagraph
    from docx.table import Table as DocxTable

    blocks: list[dict] = []
    try:
        docx_document = Document(file_path)
    except Exception as e:
        logger.warning(f"[DOCX] Gagal membuka DOCX untuk block extraction {file_path}: {e}")
        return blocks

    heading_path: list[str] = []
    block_order = 0

    # Iterasi body XML agar urutan paragraf dan tabel sesuai dokumen asli.
    # docx_document.paragraphs hanya mengembalikan paragraf, mengabaikan tabel.
    for child in docx_document.element.body:
        tag = child.tag.split("}")[-1] if "}" in child.tag else child.tag

        if tag == "p":
            # ── Paragraf ──────────────────────────────────────────────────────
            para = DocxParagraph(child, docx_document)
            text = para.text.strip()
            if not text:
                continue

            style_name = (getattr(para.style, "name", "") or "").strip()
            heading_level = None
            block_type = "paragraph"
            heading_text = heading_path[-1] if heading_path else ""

            if style_name.startswith("Heading"):
                match = re.search(r"(\d+)", style_name)
                heading_level = int(match.group(1)) if match else 1
                heading_path = heading_path[:heading_level - 1] + [text]
                block_type = "heading"
                heading_text = text
            elif _LIST_HINT_RE.match(text):
                block_type = "list_item"

            blocks.append({
                "page_number": 1,
                "text": _clean_page_text(text),
                "block_type": block_type,
                "heading_level": heading_level,
                "heading_text": heading_text,
                "heading_path": list(heading_path),
                "source_kind": "narrative",
                "block_order": block_order,
                "metadata": {"docx_style": style_name},
            })
            block_order += 1

        elif tag == "tbl":
            # ── Tabel ─────────────────────────────────────────────────────────
            try:
                table = DocxTable(child, docx_document)
                rows = table.rows
                if not rows:
                    continue

                # Baris pertama dijadikan header
                header_cells = [cell.text.strip() for cell in rows[0].cells]
                # Hapus sel duplikat dari merged cells (python-docx mengembalikan
                # sel yang sama beberapa kali untuk merged cells)
                seen: set[int] = set()
                unique_header: list[str] = []
                for cell in rows[0].cells:
                    cid = id(cell._tc)
                    if cid not in seen:
                        seen.add(cid)
                        unique_header.append(cell.text.strip())
                header_text = " | ".join(unique_header)

                current_heading = heading_path[-1] if heading_path else ""

                if len(rows) == 1:
                    # Hanya header, tidak ada baris data
                    if any(h.strip() for h in unique_header):
                        blocks.append({
                            "page_number": 1,
                            "text": header_text,
                            "block_type": "table_row",
                            "heading_level": None,
                            "heading_text": current_heading,
                            "heading_path": list(heading_path),
                            "source_kind": "table",
                            "block_order": block_order,
                            "metadata": {"is_header": True},
                        })
                        block_order += 1
                else:
                    for row_idx, row in enumerate(rows[1:], start=1):
                        seen_row: set[int] = set()
                        cells: list[str] = []
                        for cell in row.cells:
                            cid = id(cell._tc)
                            if cid not in seen_row:
                                seen_row.add(cid)
                                cells.append(cell.text.strip())
                        if not any(c for c in cells):
                            continue
                        row_text = " | ".join(cells)
                        combined = f"{header_text}\n{row_text}" if header_text else row_text
                        blocks.append({
                            "page_number": 1,
                            "text": combined,
                            "block_type": "table_row",
                            "heading_level": None,
                            "heading_text": current_heading,
                            "heading_path": list(heading_path),
                            "source_kind": "table",
                            "block_order": block_order,
                            "row_start": row_idx,
                            "row_end": row_idx,
                            "metadata": {"header_row": unique_header},
                        })
                        block_order += 1
            except Exception as e:
                logger.warning(f"[DOCX] Gagal ekstrak tabel di {file_path}: {e}")

    return blocks




def _extract_xlsx_blocks(file_path: str) -> list[dict]:
    import openpyxl

    blocks: list[dict] = []
    try:
        workbook = openpyxl.load_workbook(file_path, data_only=True, read_only=True)
    except Exception as e:
        logger.warning(f"[XLSX] Gagal membuka XLSX untuk block extraction {file_path}: {e}")
        return blocks

    block_order = 0
    for sheet in workbook.worksheets:
        header_row: list[str] = []
        row_index = 0
        blocks.append({
            "page_number": 1,
            "text": f"Sheet: {sheet.title}",
            "block_type": "sheet_header",
            "heading_level": 1,
            "heading_text": sheet.title,
            "heading_path": [sheet.title],
            "source_kind": "table",
            "sheet_name": sheet.title,
            "block_order": block_order,
            "metadata": {},
        })
        block_order += 1

        for row in sheet.iter_rows(values_only=True):
            row_index += 1
            cells = [str(cell).strip() if cell is not None else "" for cell in row]
            if not any(cells):
                continue
            if not header_row:
                header_row = cells
                blocks.append({
                    "page_number": 1,
                    "text": " | ".join(header_row),
                    "block_type": "sheet_header",
                    "heading_level": 2,
                    "heading_text": sheet.title,
                    "heading_path": [sheet.title],
                    "source_kind": "table",
                    "sheet_name": sheet.title,
                    "row_start": row_index,
                    "row_end": row_index,
                    "block_order": block_order,
                    "metadata": {"header_row": header_row},
                })
                block_order += 1
                continue

            row_text = " | ".join(cells)
            table_text = " | ".join(header_row) + "\n" + row_text
            blocks.append({
                "page_number": 1,
                "text": table_text,
                "block_type": "table_row",
                "heading_text": sheet.title,
                "heading_path": [sheet.title],
                "source_kind": "table",
                "sheet_name": sheet.title,
                "row_start": row_index,
                "row_end": row_index,
                "block_order": block_order,
                "metadata": {"header_row": header_row},
            })
            block_order += 1

    workbook.close()
    return blocks


def extract_blocks_from_file(
    file_path: str,
    lang: str = "id",
    progress_callback: Optional[Callable[..., None]] = None,
) -> list[dict]:
    """
    Extract structured blocks from supported file types.

    This complements `extract_text_from_file` for structure-aware chunking.
    """
    file_extension = os.path.splitext(file_path)[1].lower()

    if file_extension == ".pdf":
        layout_result = _extract_pdf_layout_result(
            file_path,
            dpi=config.OCR_PDF_DPI,
            retry_dpi=config.OCR_PDF_DPI_RETRY,
            progress_callback=progress_callback,
        )
        return layout_result.blocks

    if file_extension == ".docx":
        return _extract_docx_blocks(file_path)

    if file_extension == ".xlsx":
        return _extract_xlsx_blocks(file_path)

    if file_extension == ".txt":
        full_text = extract_text_from_file(
            file_path,
            lang=lang,
            return_pages=False,
            progress_callback=progress_callback,
        )
        return _build_plaintext_blocks(full_text, page_number=1, source_kind="narrative")

    if file_extension in [".jpg", ".jpeg", ".png"]:
        full_text = extract_text_from_file(
            file_path,
            lang=lang,
            return_pages=False,
            progress_callback=progress_callback,
        )
        # Mode api: hasil sudah Markdown, gunakan parser Markdown
        if config.OCR_MODE == "api":
            return _blocks_from_markdown(full_text, page_number=1, source_kind="ocr")
        return _build_plaintext_blocks(full_text, page_number=1, source_kind="ocr")

    raise ValueError(f"Format file {file_extension} belum didukung untuk block extraction.")
